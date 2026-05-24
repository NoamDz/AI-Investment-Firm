"""Tests for the third `Decisions` sheet in positions.xlsx (PLAN_reports_overhaul §T3).

Pattern lifted from ``tests/unit/test_xlsx_writer.py``; this file covers only the
new sheet so the older Positions/P&L assertions stay where they live.
"""
from __future__ import annotations

import json
from contextlib import closing
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from firm.broker.protocol import Position, Quote
from firm.db.connection import get_conn
from firm.db.migrations import init_db
from firm.reports.xlsx import write_positions_xlsx


# ---------------------------------------------------------------------------
# Stub broker — no positions, $0 cash. The Decisions sheet does not read from
# the broker so this minimal stub is sufficient.
# ---------------------------------------------------------------------------


class _EmptyBroker:
    def list_positions(self) -> list[Position]:
        return []

    def get_cash(self) -> Decimal:
        return Decimal("0.00")

    def get_quote(self, ticker: str) -> Quote:
        raise NotImplementedError

    def submit(self, decision_payload: dict[str, Any], idempotency_key: str) -> Any:  # noqa: ANN401
        raise NotImplementedError


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _as_of() -> datetime:
    return datetime(2024, 6, 1, 12, 0, 0, tzinfo=timezone.utc)


def _insert_decision(
    db: Path,
    *,
    decision_id: str,
    action: str,
    payload: dict[str, Any],
    rationale: str,
    confidence: float,
    citations: list[dict[str, Any]],
    failure_mode: str | None,
    created_at: str,
) -> None:
    row = (
        decision_id,
        json.dumps([decision_id + "-parent"]),
        action,
        json.dumps(payload),
        rationale,
        confidence,
        json.dumps(citations),
        "n/a",  # falsification
        None,   # escalation
        failure_mode,
        json.dumps({}),
        "nonce-" + decision_id,
        created_at,
    )
    with closing(get_conn(db)) as conn:
        conn.execute(
            "INSERT INTO decisions VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            row,
        )


def _seed_three_decisions(db: Path) -> None:
    """1 BUY (with citations), 1 HOLD (no failure_mode), 1 REFUSE (with failure_mode)."""
    _insert_decision(
        db,
        decision_id="dec-buy-1",
        action="BUY",
        payload={"kind": "buy", "ticker": "AAPL", "shares": "10"},
        rationale="strong momentum",
        confidence=0.85,
        citations=[{"page": 1}],
        failure_mode=None,
        created_at="2024-06-01T10:00:00+00:00",
    )
    _insert_decision(
        db,
        decision_id="dec-hold-1",
        action="HOLD",
        payload={"kind": "hold", "reason": "wait"},
        rationale="neutral",
        confidence=0.60,
        citations=[],
        failure_mode=None,
        created_at="2024-06-01T11:00:00+00:00",
    )
    _insert_decision(
        db,
        decision_id="dec-refuse-1",
        action="REFUSE",
        payload={"kind": "refuse", "reason": "no evidence"},
        rationale="insufficient evidence",
        confidence=0.30,
        citations=[],
        failure_mode="INSUFFICIENT_EVIDENCE",
        created_at="2024-06-01T11:30:00+00:00",
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_decisions_sheet_present(tmp_path: Path) -> None:
    """Workbook gains a third sheet `Decisions` with the spec'd header row."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_three_decisions(db)

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Positions", "P&L", "Decisions"]

    ws = wb["Decisions"]
    expected_headers = [
        "ts",
        "action",
        "ticker",
        "shares",
        "confidence",
        "citations",
        "failure_mode",
        "rationale",
    ]
    actual_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(expected_headers))]
    assert actual_headers == expected_headers


def test_decisions_columns_populated(tmp_path: Path) -> None:
    """Single BUY: ticker/shares/citations-count/truncated-rationale populated correctly."""
    db = tmp_path / "firm.db"
    init_db(db)
    _insert_decision(
        db,
        decision_id="dec-buy-only",
        action="BUY",
        payload={"ticker": "AAPL", "shares": "100", "kind": "buy"},
        rationale="x" * 200,
        confidence=0.91,
        citations=[{"page": 1}, {"page": 2}],
        failure_mode=None,
        created_at="2024-06-01T09:00:00+00:00",
    )

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["Decisions"]

    # Row 2 is the first (and only) data row.
    assert ws["A2"].value == "2024-06-01T09:00:00+00:00"
    assert ws["B2"].value == "BUY"
    assert ws["C2"].value == "AAPL"
    assert ws["D2"].value == pytest.approx(100.0)
    assert ws["E2"].value == pytest.approx(0.91)
    assert ws["F2"].value == 2  # citations count
    assert ws["G2"].value is None  # failure_mode
    assert ws["H2"].value == ("x" * 120) + "…"


def test_header_bold_and_frozen(tmp_path: Path) -> None:
    """Decisions header row is bold; freeze_panes locks the header in place."""
    db = tmp_path / "firm.db"
    init_db(db)
    # No decisions needed — header styling exists regardless.

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["Decisions"]

    assert ws["A1"].font.bold is True
    # Spot-check a middle and the last header cell too.
    assert ws["E1"].font.bold is True
    assert ws["H1"].font.bold is True
    assert ws.freeze_panes == "A2"


def test_holds_have_blank_ticker_shares(tmp_path: Path) -> None:
    """HOLD and REFUSE rows leave ticker and shares cells as None."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_three_decisions(db)

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["Decisions"]

    # Rows are sorted ASC: row2=BUY, row3=HOLD, row4=REFUSE.
    # Sanity-check BUY row has ticker/shares populated...
    assert ws["B2"].value == "BUY"
    assert ws["C2"].value == "AAPL"
    assert ws["D2"].value == pytest.approx(10.0)

    # HOLD row — ticker and shares must be None.
    assert ws["B3"].value == "HOLD"
    assert ws["C3"].value is None
    assert ws["D3"].value is None

    # REFUSE row — ticker and shares must be None; failure_mode populated.
    assert ws["B4"].value == "REFUSE"
    assert ws["C4"].value is None
    assert ws["D4"].value is None
    assert ws["G4"].value == "INSUFFICIENT_EVIDENCE"
