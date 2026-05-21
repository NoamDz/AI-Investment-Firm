"""Cell-by-cell tests for firm.reports.xlsx (Plan 3 §T15)."""
from __future__ import annotations

import json
from contextlib import closing
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from firm.broker.protocol import Position, Quote
from firm.db.connection import get_conn
from firm.db.migrations import init_db
from firm.reports.xlsx import write_positions_xlsx

# ---------------------------------------------------------------------------
# Stub broker
# ---------------------------------------------------------------------------


class _StubBroker:
    """Minimal Broker Protocol implementation for tests."""

    _positions = [
        Position(ticker="AAPL", shares=Decimal("10"), avg_cost=Decimal("150.00")),
        Position(ticker="MSFT", shares=Decimal("5"), avg_cost=Decimal("300.00")),
    ]
    _quotes: dict[str, Decimal] = {
        "AAPL": Decimal("170.00"),
        "MSFT": Decimal("280.00"),
    }
    _cash = Decimal("5000.00")

    def list_positions(self) -> list[Position]:
        return list(self._positions)

    def get_cash(self) -> Decimal:
        return self._cash

    def get_quote(self, ticker: str) -> Quote:
        return Quote(
            ticker=ticker,
            price=self._quotes[ticker],
            timestamp="2024-06-01T12:00:00+00:00",
        )

    def submit(self, decision_payload: dict[str, Any], idempotency_key: str) -> Any:  # noqa: ANN401
        raise NotImplementedError  # not needed in T15


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_decisions(db: Path) -> None:
    """Insert 3 decisions: BUY, SELL, HOLD."""
    rows = [
        (
            "dec-buy-1",
            json.dumps(["pm-1"]),
            "BUY",
            json.dumps({"kind": "buy", "ticker": "AAPL", "shares": "10"}),
            "strong momentum",
            0.85,
            json.dumps([]),
            "if price drops 20%",
            None,
            None,
            json.dumps({}),
            "nonce-buy-1",
            "2024-06-01T10:00:00+00:00",
        ),
        (
            "dec-sell-1",
            json.dumps(["pm-2"]),
            "SELL",
            json.dumps({"kind": "sell", "ticker": "MSFT", "shares": "5"}),
            "overvalued",
            0.75,
            json.dumps([]),
            "if PE drops below 30",
            None,
            None,
            json.dumps({}),
            "nonce-sell-1",
            "2024-06-01T11:00:00+00:00",
        ),
        (
            "dec-hold-1",
            json.dumps(["pm-3"]),
            "HOLD",
            json.dumps({"kind": "hold", "reason": "wait for earnings"}),
            "neutral",
            0.60,
            json.dumps([]),
            "if catalyst appears",
            None,
            None,
            json.dumps({}),
            "nonce-hold-1",
            "2024-06-01T11:30:00+00:00",
        ),
    ]
    with closing(get_conn(db)) as conn:
        conn.executemany(
            "INSERT INTO decisions VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )


def _as_of() -> datetime:
    return datetime(2024, 6, 1, 12, 0, 0, tzinfo=timezone.utc)


# ---------------------------------------------------------------------------
# Test: cell-by-cell Positions sheet
# ---------------------------------------------------------------------------


def test_positions_sheet_cells(tmp_path: Path) -> None:
    """Every cell in the Positions sheet matches expected computed values."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_decisions(db)

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_StubBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["Positions"]

    # Header row
    assert ws["A1"].value == "ticker"
    assert ws["B1"].value == "shares"
    assert ws["C1"].value == "avg_cost"
    assert ws["D1"].value == "current_price"
    assert ws["E1"].value == "market_value"
    assert ws["F1"].value == "unrealized_pnl"

    # AAPL row — shares=10, avg_cost=150, price=170
    # market_value = 10 * 170 = 1700; unrealized_pnl = (170-150)*10 = 200
    assert ws["A2"].value == "AAPL"
    assert ws["B2"].value == pytest.approx(10.0)
    assert ws["C2"].value == pytest.approx(150.0)
    assert ws["D2"].value == pytest.approx(170.0)
    assert ws["E2"].value == pytest.approx(1700.0)
    assert ws["F2"].value == pytest.approx(200.0)

    # MSFT row — shares=5, avg_cost=300, price=280
    # market_value = 5 * 280 = 1400; unrealized_pnl = (280-300)*5 = -100
    assert ws["A3"].value == "MSFT"
    assert ws["B3"].value == pytest.approx(5.0)
    assert ws["C3"].value == pytest.approx(300.0)
    assert ws["D3"].value == pytest.approx(280.0)
    assert ws["E3"].value == pytest.approx(1400.0)
    assert ws["F3"].value == pytest.approx(-100.0)

    # CASH row — only market_value populated
    assert ws["A4"].value == "CASH"
    assert ws["B4"].value is None
    assert ws["C4"].value is None
    assert ws["D4"].value is None
    assert ws["E4"].value == pytest.approx(5000.0)
    assert ws["F4"].value is None

    # No extra rows
    assert ws["A5"].value is None


# ---------------------------------------------------------------------------
# Test: cell-by-cell P&L sheet
# ---------------------------------------------------------------------------


def test_pnl_sheet_cells(tmp_path: Path) -> None:
    """Every cell in the P&L sheet matches seeded decisions, sorted by ts."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_decisions(db)

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_StubBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["P&L"]

    # Header
    assert ws["A1"].value == "decision_id"
    assert ws["B1"].value == "ts"
    assert ws["C1"].value == "action"
    assert ws["D1"].value == "ticker"
    assert ws["E1"].value == "shares"
    assert ws["F1"].value == "confidence"
    assert ws["G1"].value == "failure_mode"

    # Row 2 — BUY (earliest ts)
    assert ws["A2"].value == "dec-buy-1"
    assert ws["B2"].value == "2024-06-01T10:00:00+00:00"
    assert ws["C2"].value == "BUY"
    assert ws["D2"].value == "AAPL"
    assert ws["E2"].value == pytest.approx(10.0)  # numeric, not text
    assert ws["F2"].value == pytest.approx(0.85)
    assert ws["G2"].value is None  # failure_mode not set

    # Row 3 — SELL
    assert ws["A3"].value == "dec-sell-1"
    assert ws["C3"].value == "SELL"
    assert ws["D3"].value == "MSFT"
    assert ws["E3"].value == pytest.approx(5.0)  # numeric, not text

    # Row 4 — HOLD (ticker and shares are None — non-trade action)
    assert ws["A4"].value == "dec-hold-1"
    assert ws["C4"].value == "HOLD"
    assert ws["D4"].value is None
    assert ws["E4"].value is None
    assert ws["F4"].value == pytest.approx(0.60)

    # No extra rows
    assert ws["A5"].value is None


# ---------------------------------------------------------------------------
# Test: idempotency (second write overwrites cleanly)
# ---------------------------------------------------------------------------


def _dump_all_cells(path: Path) -> dict[str, list[list[Any]]]:
    """Load every cell value from every sheet as a {sheet: rows} mapping."""
    wb = openpyxl.load_workbook(path)
    return {
        sheet: [list(row) for row in wb[sheet].iter_rows(values_only=True)]
        for sheet in wb.sheetnames
    }


def test_idempotency(tmp_path: Path) -> None:
    """Calling write_positions_xlsx twice produces fully identical cell content."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_decisions(db)
    out = tmp_path / "positions.xlsx"

    write_positions_xlsx(path=out, broker=_StubBroker(), db_path=db, as_of=_as_of())
    size_first = out.stat().st_size
    cells_first = _dump_all_cells(out)

    write_positions_xlsx(path=out, broker=_StubBroker(), db_path=db, as_of=_as_of())
    size_second = out.stat().st_size
    cells_second = _dump_all_cells(out)

    assert size_first > 0
    assert size_second > 0
    # Deep cell-by-cell equality across all sheets — a regression that corrupts
    # any cell on either sheet will fail here.
    assert cells_first == cells_second


# ---------------------------------------------------------------------------
# Test: no positions, no decisions (edge case — just headers + CASH)
# ---------------------------------------------------------------------------


class _EmptyBroker:
    """Broker with no positions and zero cash."""

    def list_positions(self) -> list[Position]:
        return []

    def get_cash(self) -> Decimal:
        return Decimal("0.00")

    def get_quote(self, ticker: str) -> Quote:
        raise NotImplementedError

    def submit(self, decision_payload: dict[str, Any], idempotency_key: str) -> Any:  # noqa: ANN401
        raise NotImplementedError


def test_no_positions_no_decisions(tmp_path: Path) -> None:
    """Empty broker + empty DB produces header+CASH row and header-only P&L."""
    db = tmp_path / "firm.db"
    init_db(db)
    out = tmp_path / "positions.xlsx"

    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)

    ws_pos = wb["Positions"]
    # Header row only, then CASH
    assert ws_pos["A1"].value == "ticker"
    assert ws_pos["A2"].value == "CASH"
    assert ws_pos["E2"].value == pytest.approx(0.0)
    assert ws_pos["A3"].value is None  # nothing after CASH

    ws_pnl = wb["P&L"]
    assert ws_pnl["A1"].value == "decision_id"
    assert ws_pnl["A2"].value is None  # no decisions


# ---------------------------------------------------------------------------
# Test: parent dir created automatically
# ---------------------------------------------------------------------------


def test_creates_parent_dirs(tmp_path: Path) -> None:
    """write_positions_xlsx creates missing parent directories."""
    db = tmp_path / "firm.db"
    init_db(db)
    out = tmp_path / "nested" / "deep" / "report.xlsx"

    write_positions_xlsx(path=out, broker=_EmptyBroker(), db_path=db, as_of=_as_of())

    assert out.exists()
    assert out.stat().st_size > 0


# ---------------------------------------------------------------------------
# Test: as_of filter is strict-less-than (boundary excluded)
# ---------------------------------------------------------------------------


def test_as_of_filter_excludes_boundary(tmp_path: Path) -> None:
    """A decision with created_at == as_of must NOT appear in the P&L sheet."""
    db = tmp_path / "firm.db"
    init_db(db)
    _seed_decisions(db)

    # Seed an extra decision exactly at the as_of boundary.
    boundary_row = (
        "dec-boundary",
        json.dumps(["pm-boundary"]),
        "BUY",
        json.dumps({"kind": "buy", "ticker": "GOOG", "shares": "1"}),
        "boundary",
        0.50,
        json.dumps([]),
        "n/a",
        None,
        None,
        json.dumps({}),
        "nonce-boundary",
        "2024-06-01T12:00:00+00:00",  # == _as_of()
    )
    with closing(get_conn(db)) as conn:
        conn.execute(
            "INSERT INTO decisions VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            boundary_row,
        )

    out = tmp_path / "positions.xlsx"
    write_positions_xlsx(path=out, broker=_StubBroker(), db_path=db, as_of=_as_of())

    wb = openpyxl.load_workbook(out)
    ws = wb["P&L"]
    ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "dec-boundary" not in ids
    # The 3 pre-existing seeded decisions are all strictly before as_of, so still present.
    assert ids == ["dec-buy-1", "dec-sell-1", "dec-hold-1"]
