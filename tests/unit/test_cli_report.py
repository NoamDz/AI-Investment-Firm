"""Tests for T18: firm.cli report subcommand.

Covers:
1. test_report_writes_bundle — happy path: daily_report.md + positions.xlsx written.
2. test_report_is_idempotent — re-running overwrites; byte-for-byte md, cell-for-cell xlsx.
3. test_report_bad_date_format — non-zero exit + helpful error for invalid date.
4. test_sample_run_bundle_committed — CI invariant: all three regime sample bundles exist.
5. test_sample_run_decisions_trace_linkage — schema + trace_id linkage for all three samples.
"""
from __future__ import annotations

import json
from contextlib import closing
from decimal import Decimal
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl
import pytest

from click.testing import CliRunner

from firm.broker.protocol import Position, Quote
from firm.cli import cli
from firm.db.connection import get_conn
from firm.db.migrations import init_db


# ---------------------------------------------------------------------------
# Stub broker (no network calls)
# ---------------------------------------------------------------------------


class _StubBroker:
    """Minimal Broker Protocol implementation for tests — no external deps."""

    _positions = [
        Position(ticker="AAPL", shares=Decimal("100"), avg_cost=Decimal("150.00")),
    ]
    _quotes: dict[str, Decimal] = {
        "AAPL": Decimal("182.00"),
    }
    _cash = Decimal("94250.00")

    def list_positions(self) -> list[Position]:
        return list(self._positions)

    def get_cash(self) -> Decimal:
        return self._cash

    def get_quote(self, ticker: str) -> Quote:
        price = self._quotes.get(ticker, Decimal("100.00"))
        return Quote(
            ticker=ticker,
            price=price,
            timestamp="2024-03-13T16:00:00+00:00",
        )

    def submit(self, decision_payload: dict[str, Any], idempotency_key: str) -> Any:  # noqa: ANN401
        raise NotImplementedError  # not needed in T18 tests


# ---------------------------------------------------------------------------
# DB seeding helpers
# ---------------------------------------------------------------------------

_REPLAY_AT = "2024-03-13T16:00:00+00:00"
_REPORT_DATE = "2024-03-13"


def _seed_db(db: Path) -> None:
    """Seed a handful of decisions and a cash row for the test date."""
    ts = "2024-03-13T14:30:00+00:00"
    ts2 = "2024-03-13T15:00:00+00:00"
    with closing(get_conn(db)) as conn:
        # Cash row
        conn.execute(
            "INSERT OR REPLACE INTO cash (id, amount, updated_at) VALUES (1, ?, ?)",
            ("94250.00", ts),
        )
        # AAPL position
        conn.execute(
            "INSERT INTO positions (ticker, shares, avg_cost, updated_at) VALUES (?, ?, ?, ?)",
            ("AAPL", "100", "150.00", ts),
        )
        # Two decisions: BUY + HOLD
        decisions = [
            (
                "dec-buy-1",
                json.dumps(["pm-1"]),
                "BUY",
                json.dumps({"kind": "buy", "ticker": "AAPL", "shares": "100"}),
                "strong momentum",
                0.85,
                json.dumps([]),
                "if price drops 20%",
                None,
                None,
                json.dumps({}),
                "nonce-buy-1",
                ts,
            ),
            (
                "dec-hold-1",
                json.dumps(["pm-2"]),
                "HOLD",
                json.dumps({"kind": "hold", "reason": "wait for catalyst"}),
                "neutral stance",
                0.60,
                json.dumps([]),
                "if catalyst appears",
                None,
                None,
                json.dumps({}),
                "nonce-hold-1",
                ts2,
            ),
        ]
        conn.executemany(
            "INSERT INTO decisions VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            decisions,
        )


# ---------------------------------------------------------------------------
# Test 1 — happy path: bundle files are written
# ---------------------------------------------------------------------------


def test_report_writes_bundle(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Invoking 'report --date YYYY-MM-DD' writes daily_report.md + positions.xlsx."""
    db = tmp_path / "firm.db"
    reports_root = tmp_path / "reports"
    init_db(db)
    _seed_db(db)

    monkeypatch.setenv("FIRM_DB_PATH", str(db))
    monkeypatch.setenv("FIRM_REPORTS_ROOT", str(reports_root))
    monkeypatch.setenv("FIRM_REPLAY_AT", _REPLAY_AT)

    stub = _StubBroker()
    runner = CliRunner()
    with patch("firm.cli.make_broker", return_value=stub):
        result = runner.invoke(cli, ["report", "--date", _REPORT_DATE])

    assert result.exit_code == 0, result.output
    assert "Report bundle written" in result.output

    bundle_dir = reports_root / _REPORT_DATE
    assert (bundle_dir / "daily_report.md").exists(), "daily_report.md not found"
    assert (bundle_dir / "positions.xlsx").exists(), "positions.xlsx not found"

    # Sanity-check markdown content
    md = (bundle_dir / "daily_report.md").read_text(encoding="utf-8")
    assert "DECISION SUMMARY" in md
    assert "RECONCILIATION (EOD)" in md

    # Sanity-check xlsx is loadable and has expected sheets
    wb = openpyxl.load_workbook(bundle_dir / "positions.xlsx")
    assert "Positions" in wb.sheetnames
    assert "P&L" in wb.sheetnames


# ---------------------------------------------------------------------------
# Test 2 — idempotency: second run overwrites; files match
# ---------------------------------------------------------------------------


def test_report_is_idempotent(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Running report twice produces identical output (overwrite, not append)."""
    db = tmp_path / "firm.db"
    reports_root = tmp_path / "reports"
    init_db(db)
    _seed_db(db)

    monkeypatch.setenv("FIRM_DB_PATH", str(db))
    monkeypatch.setenv("FIRM_REPORTS_ROOT", str(reports_root))
    monkeypatch.setenv("FIRM_REPLAY_AT", _REPLAY_AT)

    stub = _StubBroker()
    runner = CliRunner()

    with patch("firm.cli.make_broker", return_value=stub):
        r1 = runner.invoke(cli, ["report", "--date", _REPORT_DATE])
        assert r1.exit_code == 0, r1.output

    bundle_dir = reports_root / _REPORT_DATE
    md_after_run1 = (bundle_dir / "daily_report.md").read_bytes()
    xlsx_after_run1 = (bundle_dir / "positions.xlsx").read_bytes()

    with patch("firm.cli.make_broker", return_value=stub):
        r2 = runner.invoke(cli, ["report", "--date", _REPORT_DATE])
        assert r2.exit_code == 0, r2.output

    md_after_run2 = (bundle_dir / "daily_report.md").read_bytes()

    # Markdown must be byte-for-byte identical.
    assert md_after_run1 == md_after_run2, "daily_report.md changed on second run"

    # xlsx: compare cell-for-cell via openpyxl (xlsx is not byte-stable due to zip timestamps).
    xlsx_after_run2 = (bundle_dir / "positions.xlsx").read_bytes()

    def _cells(wb: openpyxl.Workbook) -> list[list[Any]]:
        result: list[list[Any]] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                result.append([cell.value for cell in row])
        return result

    import io
    wb1 = openpyxl.load_workbook(io.BytesIO(xlsx_after_run1))
    wb2 = openpyxl.load_workbook(io.BytesIO(xlsx_after_run2))
    assert _cells(wb1) == _cells(wb2), "positions.xlsx cells changed on second run"


# ---------------------------------------------------------------------------
# Test 3 — bad date format: non-zero exit with useful error
# ---------------------------------------------------------------------------


def test_report_bad_date_format(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Passing --date with an invalid string must exit non-zero with an error message."""
    monkeypatch.setenv("FIRM_DB_PATH", str(tmp_path / "firm.db"))
    monkeypatch.setenv("FIRM_REPORTS_ROOT", str(tmp_path / "reports"))
    monkeypatch.setenv("FIRM_REPLAY_AT", _REPLAY_AT)

    stub = _StubBroker()
    runner = CliRunner()
    with patch("firm.cli.make_broker", return_value=stub):
        result = runner.invoke(cli, ["report", "--date", "notadate"])

    assert result.exit_code != 0, "Expected non-zero exit for bad date"
    # The error should be surfaced — either in output or exception message.
    combined = (result.output or "") + str(result.exception or "")
    assert "notadate" in combined or "YYYY-MM-DD" in combined or "date" in combined.lower(), (
        f"Expected date error message, got: {result.output!r}"
    )


# ---------------------------------------------------------------------------
# Test 4 — CI invariant: sample bundle committed
# ---------------------------------------------------------------------------


_SAMPLE_RUN_DATES: tuple[str, ...] = ("2024-03-13", "2024-08-07", "2023-11-08")


def test_sample_run_bundle_committed() -> None:
    """All sample bundle files must be committed for every regime midpoint date."""
    sample_root = Path(__file__).parent.parent.parent / "sample_runs"

    for date_str in _SAMPLE_RUN_DATES:
        sample_dir = sample_root / date_str
        assert (sample_dir / "daily_report.md").exists(), (
            f"Missing sample_runs/{date_str}/daily_report.md (sample_dir={sample_dir})"
        )
        assert (sample_dir / "positions.xlsx").exists(), (
            f"Missing sample_runs/{date_str}/positions.xlsx (sample_dir={sample_dir})"
        )
        assert (sample_dir / "decisions.jsonl").exists(), (
            f"Missing sample_runs/{date_str}/decisions.jsonl (sample_dir={sample_dir})"
        )
        assert (sample_dir / "trace.jsonl").exists(), (
            f"Missing sample_runs/{date_str}/trace.jsonl (sample_dir={sample_dir})"
        )

        wb = openpyxl.load_workbook(sample_dir / "positions.xlsx")
        assert "Positions" in wb.sheetnames, (
            f"sample_runs/{date_str}/positions.xlsx missing Positions sheet"
        )
        assert "P&L" in wb.sheetnames, (
            f"sample_runs/{date_str}/positions.xlsx missing P&L sheet"
        )


def test_sample_run_decisions_trace_linkage() -> None:
    """Every decision's trace_id must appear in the matching trace.jsonl."""
    sample_root = Path(__file__).parent.parent.parent / "sample_runs"
    required_decision_keys = {"ts", "research_decision", "trace_id"}
    required_research_keys = {
        "id",
        "action",
        "payload",
        "rationale",
        "confidence",
        "failure_mode",
    }

    for date_str in _SAMPLE_RUN_DATES:
        sample_dir = sample_root / date_str

        decisions: list[dict[str, Any]] = [
            json.loads(line)
            for line in (sample_dir / "decisions.jsonl").read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
        assert decisions, f"sample_runs/{date_str}/decisions.jsonl is empty"

        for row in decisions:
            missing = required_decision_keys - row.keys()
            assert not missing, (
                f"sample_runs/{date_str}/decisions.jsonl row missing keys {missing}"
            )
            research = row["research_decision"]
            missing_research = required_research_keys - research.keys()
            assert not missing_research, (
                f"sample_runs/{date_str}/decisions.jsonl research_decision "
                f"missing keys {missing_research}"
            )

        trace_ids: set[str] = set()
        for line in (sample_dir / "trace.jsonl").read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            span = json.loads(line)
            trace_ids.add(span["trace_id"])

        for row in decisions:
            assert row["trace_id"] in trace_ids, (
                f"sample_runs/{date_str}: decision {row['research_decision']['id']} "
                f"trace_id={row['trace_id']} not found in trace.jsonl"
            )
