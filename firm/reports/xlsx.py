"""XLSX report writer: Positions + P&L sheets. See Plan 3 §T15."""
from __future__ import annotations

import json
from contextlib import closing
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

from firm.broker.protocol import Broker
from firm.db.connection import get_conn

_POSITIONS_HEADERS = [
    "ticker",
    "shares",
    "avg_cost",
    "current_price",
    "market_value",
    "unrealized_pnl",
]

_PNL_HEADERS = [
    "decision_id",
    "ts",
    "action",
    "ticker",
    "shares",
    "confidence",
    "failure_mode",
]

_DECISIONS_HEADERS = [
    "ts",
    "action",
    "ticker",
    "shares",
    "confidence",
    "citations",
    "failure_mode",
    "rationale",
]

# Max rendered length for the rationale cell on the Decisions sheet (T3 §6).
# Longer values are truncated and suffixed with U+2026 (…).
_RATIONALE_MAX_CHARS = 120

# Actions that carry a ticker+shares in their payload.
_TRADE_ACTIONS = {"BUY", "SELL"}


def _to_float(value: Decimal) -> float:
    """Decimal → float at the cell-write boundary so spreadsheet math works."""
    return float(value)


def _extract_trade_fields(action: str, payload_json: str) -> tuple[str | None, float | None]:
    """Return (ticker, shares) for trade actions; None for non-trade actions.

    Shares is returned as float (not str) so Excel SUM/SORT on the column works —
    text-typed shares would silently no-op under spreadsheet aggregation.
    """
    if action.upper() not in _TRADE_ACTIONS:
        return None, None
    try:
        payload: dict[str, Any] = json.loads(payload_json)
        ticker_raw = payload.get("ticker")
        ticker = str(ticker_raw) if ticker_raw else None
        shares_raw = payload.get("shares")
        if shares_raw is None or shares_raw == "":
            shares: float | None = None
        else:
            shares = float(shares_raw)
        return ticker, shares
    except (json.JSONDecodeError, KeyError, ValueError):
        return None, None


def write_positions_xlsx(
    *,
    path: Path,
    broker: Broker,
    db_path: Path,
    as_of: datetime,
) -> Path:
    """Write positions.xlsx with Positions + P&L sheets. Returns path written."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Positions sheet ---
    ws_pos = wb.active
    ws_pos.title = "Positions"  # type: ignore[union-attr]
    ws_pos.append(_POSITIONS_HEADERS)  # type: ignore[union-attr]

    for pos in broker.list_positions():
        quote = broker.get_quote(pos.ticker)
        market_value = pos.shares * quote.price
        unrealized_pnl = (quote.price - pos.avg_cost) * pos.shares
        ws_pos.append(  # type: ignore[union-attr]
            [
                pos.ticker,
                _to_float(pos.shares),
                _to_float(pos.avg_cost),
                _to_float(quote.price),
                _to_float(market_value),
                _to_float(unrealized_pnl),
            ]
        )

    # CASH row: only market_value is populated; other numeric cells left blank.
    cash = broker.get_cash()
    ws_pos.append(["CASH", None, None, None, _to_float(cash), None])  # type: ignore[union-attr]

    # --- P&L sheet ---
    ws_pnl = wb.create_sheet("P&L")
    ws_pnl.append(_PNL_HEADERS)

    as_of_iso = as_of.isoformat()
    with closing(get_conn(db_path)) as conn:
        rows = conn.execute(
            "SELECT id, created_at, action, payload, confidence, failure_mode, "
            "citations, rationale "
            "FROM decisions WHERE created_at < ? ORDER BY created_at ASC",
            (as_of_iso,),
        ).fetchall()

    for row in rows:
        decision_id: str = row["id"]
        ts: str = row["created_at"]
        action: str = row["action"]
        payload_json: str = row["payload"]
        confidence: float = row["confidence"]
        failure_mode: str | None = row["failure_mode"]

        ticker, shares = _extract_trade_fields(action, payload_json)

        ws_pnl.append(
            [
                decision_id,
                ts,
                action,
                ticker,
                shares,
                confidence,
                failure_mode,  # None stays None; non-None failure_mode written as-is
            ]
        )

    # --- Decisions sheet (Plan 4 / report-overhaul §T3) ---
    # Same rows as P&L, widened with citations count + truncated rationale so the
    # spreadsheet is self-contained if forwarded without the HTML report.
    ws_dec = wb.create_sheet("Decisions")
    ws_dec.append(_DECISIONS_HEADERS)
    for col_idx in range(1, len(_DECISIONS_HEADERS) + 1):
        ws_dec.cell(row=1, column=col_idx).font = Font(bold=True)
    ws_dec.freeze_panes = "A2"

    for row in rows:
        ts = row["created_at"]
        action = row["action"]
        payload_json = row["payload"]
        confidence = row["confidence"]
        failure_mode = row["failure_mode"]
        citations_json: str | None = row["citations"]
        rationale: str = row["rationale"] or ""

        ticker, shares = _extract_trade_fields(action, payload_json)

        try:
            citations_count = len(json.loads(citations_json)) if citations_json else 0
        except (json.JSONDecodeError, TypeError):
            citations_count = 0

        if len(rationale) > _RATIONALE_MAX_CHARS:
            rationale_cell = rationale[:_RATIONALE_MAX_CHARS] + "…"
        else:
            rationale_cell = rationale

        ws_dec.append(
            [
                ts,
                action,
                ticker,
                shares,
                confidence,
                citations_count,
                failure_mode,
                rationale_cell,
            ]
        )

    wb.save(path)
    return path
